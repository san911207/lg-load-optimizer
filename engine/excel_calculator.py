"""
Quick Fit Calculator — standalone Excel workbook generator.

US units throughout (in / lb / ft / ft³). Same closed-form fit formula as
engine.best_packer.fits_formula() exposed as Excel cells so a planner can drop
in SKUs and instantly see whether the load fits — no Streamlit, no Python.

Layout:
  Sheet "Quick Fit Check"  — user inputs + auto-calc + verdict + utilization
  Sheet "Model_Master"     — reference SKU specs (w/d/h in inches, weight in lb)
  Sheet "Truck_Master"     — reference truck specs (in / lb / ft³)
"""
from pathlib import Path
from typing import Any, Dict

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule, CellIsRule

DOOR_TRACK_LOSS_IN = 10
N_INPUT_ROWS = 30

# Headers for the input table
HEADERS = [
    "#", "Model Code", "Qty",
    "w (in)", "d (in)", "h (in)", "Stackable",
    "Layers", "Length A (ft)", "Length B (ft)",
    "Min Length (ft)", "Volume (ft³)", "Weight (lb)",
]

# CLAUDE.md calibrations (US units)
CALIBRATIONS = {
    "LDFN4542S": {"stackable": True, "load_bear_lb": 132.3, "fragile": False},
    "LWS3063ST": {"stackable": True, "load_bear_lb": 198.4, "fragile": False},
}

# Sample pre-fill (L001)
SAMPLE_L001 = [
    ("LF29H8330S", 6),
    ("WM4000HWA", 8),
    ("DLEX4000W", 8),
    ("LDFN4542S", 10),
    ("LMV1764ST", 12),
]

# Color helpers
GREEN = "1D9E75"
ORANGE = "E89F32"
GRAY = "9CA3AF"
RED = "DC2626"
LIGHT_GREEN_BG = "F0FDF4"
LIGHT_BLUE_BG = "EFF6FF"
HEADER_BG = "0F6E56"


def build_fit_calculator(master_xlsx_path: Path, out_path: Path) -> Path:
    master_df = pd.read_excel(master_xlsx_path, sheet_name="Model_Master")
    truck_df = pd.read_excel(master_xlsx_path, sheet_name="Truck_Master")

    # Calibrations
    for mc, vals in CALIBRATIONS.items():
        mask = master_df["model_code"] == mc
        for col, v in vals.items():
            master_df.loc[mask, col] = v

    wb = Workbook()
    ws_main = wb.active
    ws_main.title = "Quick Fit Check"
    _build_main_sheet(ws_main, master_df)

    ws_master = wb.create_sheet("Model_Master")
    _write_df_sheet(ws_master, master_df)

    ws_trucks = wb.create_sheet("Truck_Master")
    _write_df_sheet(ws_trucks, truck_df)

    wb.save(out_path)
    return out_path


def _build_main_sheet(ws, master_df: pd.DataFrame) -> None:
    # Title
    ws["A1"] = "LG Quick Fit Calculator"
    ws["A1"].font = Font(size=18, bold=True, color="0F6E56")
    ws.merge_cells("A1:M1")
    ws["A2"] = (
        "Closed-form fit predictor — all US units (in / lb / ft³). "
        "Max-fit mode: load_bear / fragile ignored, door track (10 in) auto-applied."
    )
    ws["A2"].font = Font(italic=True, color="6B7280")
    ws.merge_cells("A2:M2")

    # Instructions
    ws["A4"] = "How to use"
    ws["A4"].font = Font(bold=True, size=12)
    ws["A5"] = "1. Pick truck (26ft or 53ft) in cell C9 (dropdown)."
    ws["A6"] = "2. For each item: enter SKU code in column B (dropdown) and quantity in column C."
    ws["A7"] = (
        "3. Same-dim SKUs (washer + dryer, etc.): combine into ONE row "
        "for the simulator-exact answer."
    )

    # Truck selector + specs
    ws["B9"] = "Truck:"
    ws["B9"].font = Font(bold=True)
    ws["C9"] = "26ft"
    ws["C9"].font = Font(bold=True, size=13, color="0F6E56")
    ws["C9"].fill = PatternFill("solid", fgColor=LIGHT_GREEN_BG)
    truck_dv = DataValidation(type="list", formula1='"26ft,53ft"', allow_blank=False)
    truck_dv.add("C9")
    ws.add_data_validation(truck_dv)

    # Truck dims — stored as inches for math, also shown in feet
    ws["B10"] = "Truck length (in):"
    ws["C10"] = "=VLOOKUP(C9,Truck_Master!A:G,3,FALSE)"
    ws["C10"].number_format = "0.0"
    ws["D10"] = "=C10/12"
    ws["D10"].number_format = "0.00\" ft\""

    ws["B11"] = "Truck width (in):"
    ws["C11"] = "=VLOOKUP(C9,Truck_Master!A:G,4,FALSE)"
    ws["C11"].number_format = "0.0"
    ws["D11"] = "=C11/12"
    ws["D11"].number_format = "0.00\" ft\""

    ws["B12"] = "Truck height (in):"
    ws["C12"] = "=VLOOKUP(C9,Truck_Master!A:G,5,FALSE)"
    ws["C12"].number_format = "0.0"
    ws["D12"] = "=C12/12"
    ws["D12"].number_format = "0.00\" ft\""

    ws["B13"] = "Effective height (in):"
    ws["C13"] = "=C12-10"
    ws["C13"].number_format = "0.0"
    ws["D13"] = "(after 10\" door track loss)"
    ws["D13"].font = Font(italic=True, color="6B7280")

    ws["B14"] = "Max payload (lb):"
    ws["C14"] = "=VLOOKUP(C9,Truck_Master!A:G,6,FALSE)"
    ws["C14"].number_format = "#,##0"

    ws["B15"] = "Cargo volume (ft³):"
    ws["C15"] = "=VLOOKUP(C9,Truck_Master!A:G,7,FALSE)"
    ws["C15"].number_format = "#,##0.0"

    # Input table header
    header_row = 17
    for i, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=header_row, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=HEADER_BG)
        c.alignment = Alignment(horizontal="center")

    first_input = header_row + 1
    last_input = first_input + N_INPUT_ROWS - 1

    model_codes = master_df["model_code"].tolist()
    model_dv = DataValidation(
        type="list",
        formula1=f"=Model_Master!$A$2:$A${len(model_codes) + 1}",
        allow_blank=True,
    )

    # Master columns (1-indexed): 1=code 2=cat 3=w_in 4=d_in 5=h_in 6=weight_lb
    #                             7=this_side_up 8=stackable 9=load_bear_lb
    #                             10=fragile 11=notes 12=volume_cft
    for r in range(first_input, last_input + 1):
        ws.cell(row=r, column=1, value=r - header_row)
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=1).font = Font(color="9CA3AF")
        for col in (2, 3):
            ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=LIGHT_BLUE_BG)
        model_dv.add(f"B{r}")

        w_in = f'VLOOKUP(B{r},Model_Master!A:L,3,FALSE)'
        d_in = f'VLOOKUP(B{r},Model_Master!A:L,4,FALSE)'
        h_in = f'VLOOKUP(B{r},Model_Master!A:L,5,FALSE)'
        weight_lb = f'VLOOKUP(B{r},Model_Master!A:L,6,FALSE)'
        stack = f'VLOOKUP(B{r},Model_Master!A:L,8,FALSE)'
        vol_cft = f'VLOOKUP(B{r},Model_Master!A:L,12,FALSE)'

        # D: w (in), E: d (in), F: h (in) — direct lookup, no conversion
        ws.cell(row=r, column=4, value=f'=IF(B{r}="","",{w_in})')
        ws.cell(row=r, column=4).number_format = "0.0"
        ws.cell(row=r, column=5, value=f'=IF(B{r}="","",{d_in})')
        ws.cell(row=r, column=5).number_format = "0.0"
        ws.cell(row=r, column=6, value=f'=IF(B{r}="","",{h_in})')
        ws.cell(row=r, column=6).number_format = "0.0"

        # G: stackable
        ws.cell(row=r, column=7, value=f'=IF(B{r}="","",{stack})')

        # H: layers = floor(eff_H / h_in) if stackable else 1
        ws.cell(row=r, column=8, value=(
            f'=IF(B{r}="","",IF({stack}=TRUE,FLOOR($C$13/F{r},1),1))'
        ))

        # I: Length A (ft) — orient A
        ws.cell(row=r, column=9, value=(
            f'=IF(B{r}="","",CEILING(C{r}/(FLOOR($C$11/D{r},1)*H{r}),1)*E{r}/12)'
        ))
        ws.cell(row=r, column=9).number_format = "0.00"

        # J: Length B (ft) — orient B (90° rotated)
        ws.cell(row=r, column=10, value=(
            f'=IF(B{r}="","",CEILING(C{r}/(FLOOR($C$11/E{r},1)*H{r}),1)*D{r}/12)'
        ))
        ws.cell(row=r, column=10).number_format = "0.00"

        # K: Min Length (ft)
        ws.cell(row=r, column=11, value=f'=IF(B{r}="","",MIN(I{r},J{r}))')
        ws.cell(row=r, column=11).number_format = "0.00"

        # L: Volume per row (ft³) = qty × volume_cft (direct from master col 12)
        ws.cell(row=r, column=12, value=(
            f'=IF(B{r}="","",C{r}*{vol_cft})'
        ))
        ws.cell(row=r, column=12).number_format = "0.0"

        # M: Weight per row (lb) = qty × weight_lb (already in pounds)
        ws.cell(row=r, column=13, value=(
            f'=IF(B{r}="","",C{r}*{weight_lb})'
        ))
        ws.cell(row=r, column=13).number_format = "#,##0"

    ws.add_data_validation(model_dv)

    # Pre-fill L001 sample
    for i, (mc, qty) in enumerate(SAMPLE_L001):
        ws.cell(row=first_input + i, column=2, value=mc)
        ws.cell(row=first_input + i, column=3, value=qty)

    # Totals + Result block
    total_row = last_input + 2
    truck_row = total_row + 1
    util_row = total_row + 2
    verdict_row = total_row + 3
    margin_row = total_row + 4

    # Σ row
    ws.cell(row=total_row, column=8, value="Σ Total:").font = Font(bold=True)
    ws.cell(row=total_row, column=8).alignment = Alignment(horizontal="right")

    sum_length = ws.cell(row=total_row, column=11,
                         value=f"=SUM(K{first_input}:K{last_input})")
    sum_length.font = Font(bold=True, size=12)
    sum_length.number_format = "0.00\" ft\""

    sum_vol = ws.cell(row=total_row, column=12,
                      value=f"=SUM(L{first_input}:L{last_input})")
    sum_vol.font = Font(bold=True, size=12)
    sum_vol.number_format = "0.0\" ft³\""

    sum_wt = ws.cell(row=total_row, column=13,
                     value=f"=SUM(M{first_input}:M{last_input})")
    sum_wt.font = Font(bold=True, size=12)
    sum_wt.number_format = "#,##0\" lb\""

    # Truck capacity row
    ws.cell(row=truck_row, column=8, value="Truck capacity:").font = Font(bold=True)
    ws.cell(row=truck_row, column=8).alignment = Alignment(horizontal="right")
    ws.cell(row=truck_row, column=11, value="=C10/12")
    ws.cell(row=truck_row, column=11).number_format = "0.00\" ft\""
    ws.cell(row=truck_row, column=11).font = Font(bold=True)
    ws.cell(row=truck_row, column=12, value="=C15")
    ws.cell(row=truck_row, column=12).number_format = "#,##0.0\" ft³\""
    ws.cell(row=truck_row, column=12).font = Font(bold=True)
    ws.cell(row=truck_row, column=13, value="=C14")
    ws.cell(row=truck_row, column=13).number_format = "#,##0\" lb\""
    ws.cell(row=truck_row, column=13).font = Font(bold=True)

    # Utilization % row — Σ / truck × 100
    ws.cell(row=util_row, column=8, value="Utilization:").font = Font(bold=True)
    ws.cell(row=util_row, column=8).alignment = Alignment(horizontal="right")

    util_len = ws.cell(row=util_row, column=11,
                       value=f"=K{total_row}/K{truck_row}*100")
    util_len.number_format = "0.0\" %\""
    util_len.font = Font(bold=True, size=12, color="FFFFFF")
    util_len.alignment = Alignment(horizontal="center")

    util_vol = ws.cell(row=util_row, column=12,
                       value=f"=L{total_row}/L{truck_row}*100")
    util_vol.number_format = "0.0\" %\""
    util_vol.font = Font(bold=True, size=12, color="FFFFFF")
    util_vol.alignment = Alignment(horizontal="center")

    util_wt = ws.cell(row=util_row, column=13,
                      value=f"=M{total_row}/M{truck_row}*100")
    util_wt.number_format = "0.0\" %\""
    util_wt.font = Font(bold=True, size=12, color="FFFFFF")
    util_wt.alignment = Alignment(horizontal="center")

    # Color util cells: green ≥70, orange 40-69, gray <40
    for col_letter in ("K", "L", "M"):
        cell_ref = f"{col_letter}{util_row}"
        ws.conditional_formatting.add(
            cell_ref,
            CellIsRule(operator="greaterThanOrEqual", formula=["70"],
                       fill=PatternFill("solid", fgColor=GREEN)),
        )
        ws.conditional_formatting.add(
            cell_ref,
            CellIsRule(operator="between", formula=["40", "69.999"],
                       fill=PatternFill("solid", fgColor=ORANGE)),
        )
        ws.conditional_formatting.add(
            cell_ref,
            CellIsRule(operator="lessThan", formula=["40"],
                       fill=PatternFill("solid", fgColor=GRAY)),
        )

    # Verdict row — length AND payload AND volume must all pass
    ws.cell(row=verdict_row, column=8, value="Verdict:").font = Font(bold=True, size=14)
    ws.cell(row=verdict_row, column=8).alignment = Alignment(horizontal="right")
    verdict_cell = ws.cell(
        row=verdict_row, column=11,
        value=(
            f'=IF(AND(K{total_row}<=K{truck_row},'
            f'L{total_row}<=L{truck_row},'
            f'M{total_row}<=M{truck_row}),'
            f'"✓ FITS","✗ DOES NOT FIT")'
        ),
    )
    verdict_cell.font = Font(bold=True, size=14, color="FFFFFF")
    verdict_cell.alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=verdict_row, start_column=11,
                   end_row=verdict_row, end_column=13)

    ws.conditional_formatting.add(
        f"K{verdict_row}:M{verdict_row}",
        FormulaRule(formula=[f'$K${verdict_row}="✓ FITS"'],
                    fill=PatternFill("solid", fgColor=GREEN)),
    )
    ws.conditional_formatting.add(
        f"K{verdict_row}:M{verdict_row}",
        FormulaRule(formula=[f'$K${verdict_row}="✗ DOES NOT FIT"'],
                    fill=PatternFill("solid", fgColor=RED)),
    )

    # Margin row — length / volume / weight remaining (negative = over)
    ws.cell(row=margin_row, column=8, value="Margin:").font = Font(bold=True)
    ws.cell(row=margin_row, column=8).alignment = Alignment(horizontal="right")
    ws.cell(row=margin_row, column=11, value=f"=K{truck_row}-K{total_row}")
    ws.cell(row=margin_row, column=11).number_format = '+0.00" ft";-0.00" ft"'
    ws.cell(row=margin_row, column=12, value=f"=L{truck_row}-L{total_row}")
    ws.cell(row=margin_row, column=12).number_format = '+0.0" ft³";-0.0" ft³"'
    ws.cell(row=margin_row, column=13, value=f"=M{truck_row}-M{total_row}")
    ws.cell(row=margin_row, column=13).number_format = '+#,##0" lb";-#,##0" lb"'

    # Column widths + freeze
    widths = {"A": 5, "B": 16, "C": 9, "D": 9, "E": 9, "F": 9,
              "G": 11, "H": 9, "I": 14, "J": 14, "K": 16, "L": 14, "M": 14}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A18"


def _write_df_sheet(ws, df: pd.DataFrame) -> None:
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="374151")
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_len = max((len(str(c.value or "")) for c in column_cells), default=10)
        ws.column_dimensions[
            ws.cell(row=1, column=col_idx).column_letter
        ].width = min(max_len + 3, 30)


if __name__ == "__main__":
    proj_dir = Path(__file__).resolve().parent.parent
    src = proj_dir / "data" / "sample_input.xlsx"
    out = proj_dir / "outputs" / "LG_Quick_Fit_Calculator.xlsx"
    out.parent.mkdir(exist_ok=True)
    build_fit_calculator(src, out)
    print(f"Built: {out}")
