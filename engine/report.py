"""
Excel 리포트 생성기 - 모든 Load 결과를 1개 Excel에 통합
- Summary 시트: 모든 Load 한눈에
- L001, L002, ... 각 Load별 상세 시트
- Unfitted 시트: 미적재 통합
"""
from pathlib import Path
from typing import List
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from engine.packer import PackResult


def write_excel_report(results: List[PackResult], out_path: Path):
    rows_summary = []
    rows_unfitted_all = []

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        # 1) Summary 시트
        for res in results:
            rows_summary.append({
                "Load_ID": res.load_id,
                "Truck": res.truck_display,
                "Truck_Volume_CBM": round(res.truck_volume_cbm, 2),
                "Used_Volume_CBM": res.used_volume_cbm,
                "Volume_Util_%": res.volume_util_pct,
                "Max_Payload_KG": res.truck_max_payload_kg,
                "Used_Weight_KG": res.used_weight_kg,
                "Weight_Util_%": res.weight_util_pct,
                "Fitted_EA": res.fitted_count,
                "Unfitted_EA": res.unfitted_count,
                "Status": "✅ OK" if res.unfitted_count == 0 else "⚠️ 추가트럭 필요",
            })
        df_sum = pd.DataFrame(rows_summary)
        df_sum.to_excel(writer, sheet_name="Summary", index=False)

        # 2) Load별 상세 시트
        for res in results:
            rows = []
            for it in res.fitted_items:
                rows.append({
                    "Seq": it.seq,
                    "Zone": it.zone,
                    "Model_Code": it.model_code,
                    "Category": it.category,
                    "Pos_X_mm": int(it.pos_x),
                    "Pos_Y_mm": int(it.pos_y),
                    "Pos_Z_mm": int(it.pos_z),
                    "Dim_X_mm": int(it.dim_x),
                    "Dim_Y_mm": int(it.dim_y),
                    "Dim_Z_mm": int(it.dim_z),
                    "Weight_KG": round(it.weight_kg, 1),
                    "Rotation": it.rotation,
                })
            df_detail = pd.DataFrame(rows)
            sheet = f"{res.load_id}_LoadSheet"[:31]
            df_detail.to_excel(writer, sheet_name=sheet, index=False)

            # Unfitted 누적
            for u in res.unfitted_items:
                rows_unfitted_all.append({
                    "Load_ID": res.load_id,
                    "Model_Code": u["model_code"],
                    "Quantity": u["quantity"],
                })

        # 3) Unfitted 통합
        if rows_unfitted_all:
            df_unfit = pd.DataFrame(rows_unfitted_all)
            df_unfit.to_excel(writer, sheet_name="Unfitted_All", index=False)
        else:
            pd.DataFrame([{"Note": "All items fitted ✅"}]).to_excel(
                writer, sheet_name="Unfitted_All", index=False)

    # 4) 서식 다듬기 (Summary 강조)
    from openpyxl import load_workbook
    wb = load_workbook(out_path)
    ws = wb["Summary"]
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    # 칼럼 너비 자동
    for col_idx, col in enumerate(ws.columns, 1):
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 3
    wb.save(out_path)
    return out_path
